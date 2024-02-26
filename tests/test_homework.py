from zipfile import ZipFile
from pypdf import PdfReader
import csv
from openpyxl import load_workbook

from tests.conftest import resource_zip_path


def test_pdf():
    with ZipFile(resource_zip_path) as zip_file:
        with zip_file.open('PDF.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            number_of_pages = len(reader.pages)
            assert number_of_pages == 4


def test_csv():
    with ZipFile(resource_zip_path) as zip_file:
        with zip_file.open('CSV.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            reader = list(csv.reader(content.splitlines()))
            second_row = reader[1]
            assert 'Dett' in second_row
            assert len(second_row) == 6


def test_xlsx():
    with ZipFile(resource_zip_path) as zip_file:
        with zip_file.open('XLSX.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            name = sheet.cell(row=4, column=2).value
            assert name == 'Kallsie'
